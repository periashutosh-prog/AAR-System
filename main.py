import asyncio
import io
import json
import uuid
from datetime import datetime, timezone
from urllib.parse import quote

import openpyxl
from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from supabase import Client, create_client
from supabase_auth.errors import AuthApiError

# ---------- Favicon ----------
# Gradient square + checkmark, matching the accent colors used everywhere
# else - inlined as a data URI so no static file route is needed.
FAVICON_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32">
<defs><linearGradient id="g" x1="0" y1="0" x2="1" y2="1">
<stop offset="0" stop-color="#4f8cff"/><stop offset="1" stop-color="#5fd0c5"/>
</linearGradient></defs>
<rect width="32" height="32" rx="8" fill="url(#g)"/>
<path d="M9 17l5 5 9-11" stroke="#0b0d12" stroke-width="3.4" fill="none" stroke-linecap="round" stroke-linejoin="round"/>
</svg>"""
FAVICON_DATA_URI = "data:image/svg+xml," + quote(FAVICON_SVG)

# ---------- Database ----------
# Private repo only - creds hardcoded on purpose, no env var indirection.
# Create both tables in the Supabase SQL editor first:
#
#   create table profiles (
#     id uuid primary key default gen_random_uuid(),
#     first_name text not null,
#     last_name text not null,
#     institute_name text not null,
#     institute_branch_name text,
#     email text not null,
#     mobile_number text not null,
#     student_count_range text not null
#       check (student_count_range in ('<250', '250-500', '500-1000', '1000-3000', '3000+')),
#     created_at timestamptz not null default now()
#   );
#
#   create table students (
#     aaruid text primary key,
#     institute_id uuid not null references profiles(id),
#     rfid text not null,
#     name text not null,
#     class text not null,
#     section text not null
#   );
#
#   create table devices (
#     aaruid text primary key,
#     institute_id uuid not null references profiles(id),
#     mac_address text not null,
#     name text not null,
#     class text not null,
#     section text not null
#   );
SUPABASE_URL = "https://iceeivfnpgmjdrmbxjkp.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImljZWVpdmZucGdtamRybWJ4amtwIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODM4NDQzNTEsImV4cCI6MjA5OTQyMDM1MX0._fNQ7F-A9M3-PVG7FeGx-jwcN4ZrufOXFq1DCZpzAFw"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
STUDENTS_TABLE = "students"
PROFILES_TABLE = "profiles"
DEVICES_TABLE = "devices"
IMAGES_BUCKET = "IMAGES"
STUDENT_COUNT_RANGES = ["<250", "250-500", "500-1000", "1000-3000", "3000+"]

# Excel imports are a two-step flow (analyze, then commit) so the confirmation
# step doesn't need the file re-uploaded. Parsed rows sit here in between,
# keyed by a short-lived token. Fine for a single-process dev deployment.
PENDING_IMPORTS: dict[str, dict] = {}


def generate_aaruid() -> str:
    return "AAR-" + uuid.uuid4().hex[:8].upper()


async def get_institute_id(authorization: str | None = Header(None)) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid Authorization header")
    token = authorization.removeprefix("Bearer ")
    try:
        user_response = await asyncio.to_thread(supabase.auth.get_user, token)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    if not user_response or not user_response.user:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return user_response.user.id


# ---------- WebSocket connection registry ----------
class DeviceRegistry:
    def __init__(self) -> None:
        self.connections: set[WebSocket] = set()

    async def register(self, ws: WebSocket) -> None:
        await ws.accept()
        self.connections.add(ws)

    def unregister(self, ws: WebSocket) -> None:
        self.connections.discard(ws)

    async def broadcast(self, message: dict, exclude: WebSocket | None = None) -> None:
        payload = json.dumps(message)
        dead: list[WebSocket] = []
        for ws in self.connections:
            if ws is exclude:
                continue
            try:
                await ws.send_text(payload)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.unregister(ws)


registry = DeviceRegistry()

app = FastAPI(title="Attendance System Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- Shared page shell ----------
BASE_STYLE = """
  :root{
    --bg-0:#0b0d12; --bg-1:#12151d; --card:#181c26; --card-2:#1f2430;
    --line:#2a3040; --text:#eaedf3; --muted:#8b93a6;
    --accent:#4f8cff; --accent-2:#5fd0c5;
    --present:#3ddc97; --absent:#ff5c6a; --onduty:#f7c948;
  }
  *{ box-sizing:border-box; }
  body{
    margin:0;
    min-height:100vh;
    background:
      radial-gradient(120% 60% at 50% -10%, #1a2135 0%, transparent 60%),
      linear-gradient(180deg, var(--bg-0), var(--bg-1));
    color:var(--text);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
  }
  nav{
    display:flex; align-items:center; justify-content:space-between;
    padding: 1.1rem clamp(1.25rem, 5vw, 3rem);
    border-bottom: 1px solid var(--line);
  }
  .brand{ display:flex; flex-direction:column; align-items:center; text-align:center; gap:0.15rem; line-height:1; text-decoration:none; }
  .brand-mark{ font-weight:900; letter-spacing:0.02em; font-size:2.2rem; color:var(--accent); }
  .brand-full{
    font-size:0.62rem; font-weight:700; letter-spacing:0.1em; text-transform:uppercase;
    color: var(--muted);
  }
  .nav-links{ display:flex; align-items:center; gap: 0.9rem; }
  .btn{
    display:inline-flex; align-items:center; justify-content:center;
    font-weight:600; font-size:0.9rem; text-decoration:none;
    padding: 0.55rem 1.1rem; border-radius: 10px;
    border: 1px solid var(--line);
    transition: transform 0.12s ease, border-color 0.2s ease;
  }
  .btn:active{ transform: scale(0.97); }
  .btn:disabled{ opacity:0.6; cursor:not-allowed; transform:none; }
  .btn-ghost{ color: var(--text); background: transparent; }
  .btn-ghost:hover{ border-color: var(--accent); }
  .btn-primary{
    color:#0b0d12; border-color: transparent;
    background: linear-gradient(135deg, var(--accent), var(--accent-2));
  }

  /* ---------- Shared modal overlay ---------- */
  .overlay{
    position:fixed; inset:0; background: rgba(6,7,10,0.72); backdrop-filter: blur(4px);
    display:flex; align-items:center; justify-content:center; padding: 1.25rem;
    opacity:0; pointer-events:none; transition: opacity 0.2s ease; z-index:80;
  }
  .overlay.show{ opacity:1; pointer-events:auto; }
  .overlay .modal{
    width:100%; max-width: 24rem;
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line); border-radius: 22px;
    padding: 2rem 1.8rem 1.8rem; text-align:center;
    transform: translateY(0.6rem) scale(0.98);
    transition: transform 0.22s ease;
  }
  .overlay.show .modal{ transform: translateY(0) scale(1); }
  .modal-icon{
    width:3.6rem; height:3.6rem; border-radius:50%; margin:0 auto 1.1rem;
    display:flex; align-items:center; justify-content:center;
    font-size:1.7rem; font-weight:800;
  }
  .modal-icon.danger{ background: rgba(255,92,106,0.15); color: var(--absent); border: 2px solid rgba(255,92,106,0.4); }
  .modal-icon.warn{ background: rgba(247,201,72,0.15); color: var(--onduty); border: 2px solid rgba(247,201,72,0.4); }
  .modal-icon.present{ background: rgba(61,220,151,0.15); color: var(--present); border: 2px solid rgba(61,220,151,0.4); }
  .modal h3{ margin:0 0 0.5rem; font-size:1.12rem; }
  .modal p{ color: var(--muted); font-size:0.9rem; line-height:1.55; margin:0 0 1.5rem; }
  .modal .btn{ width:100%; }
  .modal input{
    width:100%; background: rgba(255,255,255,0.04);
    border: 1px solid var(--line); border-radius: 10px;
    padding: 0.7rem 0.8rem; color: var(--text); font-size: 0.95rem;
    outline:none; font-family: inherit; margin-bottom: 0.9rem;
    transition: border-color 0.15s ease, box-shadow 0.15s ease;
  }
  .modal input:focus{ border-color: var(--accent); box-shadow: 0 0 0 3px rgba(79,140,255,0.16); }
  .modal-note{ color: var(--muted); font-size: 0.8rem; margin-top: 0.8rem; min-height:1.1rem; }

  /* ---------- Shared form card (sign up / sign in) ---------- */
  .form-wrap{
    display:flex; justify-content:center;
    padding: clamp(2rem, 8vh, 4rem) 1.25rem;
  }
  .form-card{
    width:100%; max-width: 32rem;
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line);
    border-radius: 24px;
    padding: clamp(1.8rem, 5vw, 2.6rem);
    box-shadow: 0 0.6rem 1.6rem rgba(0,0,0,0.35);
  }
  .form-card h1{ font-size: clamp(1.5rem, 4vw, 1.9rem); margin: 0 0 0.4rem; text-align:center; }
  .form-card .form-sub{ color: var(--muted); font-size: 0.9rem; text-align:center; margin: 0 0 1.8rem; }
  .field-label{
    font-size: 0.72rem; text-transform:uppercase; letter-spacing:0.06em;
    color: var(--muted); margin: 1rem 0 0.35rem;
  }
  form > .field-label:first-child,
  form > .name-row:first-child .field-label{ margin-top:0; }
  .form-card input, .form-card select{
    width:100%; background: rgba(255,255,255,0.04);
    border: 1px solid var(--line); border-radius: 10px;
    padding: 0.7rem 0.8rem; color: var(--text); font-size: 0.95rem;
    outline:none; font-family: inherit;
    transition: border-color 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
  }
  .form-card input:hover, .form-card select:hover{ border-color: rgba(79,140,255,0.45); }
  .form-card input:focus, .form-card select:focus{
    border-color: var(--accent);
    box-shadow: 0 0 0 3px rgba(79,140,255,0.16);
    background: rgba(79,140,255,0.05);
  }
  .optional{ color: var(--muted); text-transform:none; letter-spacing:0; font-weight:400; }
  .form-submit{
    width:100%; margin-top: 1.6rem; padding: 0.8rem 1rem;
    font-size: 0.98rem;
  }
  .form-note{
    text-align:center; color: var(--muted); font-size: 0.8rem;
    margin-top: 0.9rem; min-height: 1.2rem;
  }
  .form-links{ display:flex; justify-content:flex-end; margin: 0.6rem 0 0.2rem; }
  .form-links button{
    background:none; border:none; color: var(--accent-2);
    font-size: 0.82rem; font-weight:600; cursor:pointer; padding:0.2rem 0;
  }
  .form-links button:hover{ text-decoration:underline; }
  .form-switch{ text-align:center; color: var(--muted); font-size: 0.85rem; margin-top: 1.4rem; }
  .form-switch a{ color: var(--accent-2); font-weight:600; text-decoration:none; }
  .form-switch a:hover{ text-decoration:underline; }
"""


def page_shell(body: str, title: str = "AAR") -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<link rel="icon" type="image/svg+xml" href="{FAVICON_DATA_URI}">
<style>{BASE_STYLE}</style>
</head>
<body>
{body}
</body>
</html>"""


NAV = """
<nav>
  <a class="brand" href="/">
    <span class="brand-mark">AAR</span>
    <span class="brand-full">Automatic Attendance Registrar</span>
  </a>
  <div class="nav-links">
    <a class="btn btn-ghost" href="/signin" id="navSignIn">Sign In</a>
    <a class="btn btn-primary" href="/signup" id="navSignUp">Sign Up</a>
    <button type="button" class="btn btn-ghost" id="navLogOut" style="display:none;">Log Out</button>
  </div>
</nav>
<script>
  (function(){
    var loggedIn = !!localStorage.getItem('aar_access_token');
    var path = window.location.pathname;

    if (loggedIn && (path === '/signin' || path === '/signup')){
      window.location.href = '/dashboard';
      return;
    }

    var signInLink = document.getElementById('navSignIn');
    var signUpLink = document.getElementById('navSignUp');
    var logOutBtn = document.getElementById('navLogOut');
    if (loggedIn){
      signInLink.style.display = 'none';
      signUpLink.style.display = 'none';
      logOutBtn.style.display = 'inline-flex';
    }
    logOutBtn.addEventListener('click', function(){
      localStorage.removeItem('aar_access_token');
      localStorage.removeItem('aar_refresh_token');
      window.location.href = '/signin';
    });
  })();
</script>
"""


# ---------- Landing page ----------
LANDING_PAGE = page_shell(
    NAV + """
<style>
  .page{
    width:100%; max-width: 56rem; margin: 0 auto;
    padding: clamp(2rem, 8vh, 4rem) 1.25rem;
    display:flex; flex-direction:column; align-items:center; gap: clamp(2.5rem, 6vw, 4rem);
  }
  .hero-card{
    width:100%; text-align:center;
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line);
    border-radius: 28px;
    padding: clamp(2rem, 6vw, 3rem) clamp(1.4rem, 5vw, 2.4rem);
    box-shadow: 0 0.6rem 1.6rem rgba(0,0,0,0.35);
  }
  .kicker{
    color: var(--accent-2); font-weight:700; font-size:0.95rem;
    margin: 0 0 0.6rem; letter-spacing:0.03em;
  }
  .hero-card h1{
    font-size: clamp(2.4rem, 8vw, 3.6rem); margin: 0 0 0.7rem; letter-spacing: -0.02em;
    background: linear-gradient(135deg, var(--accent), var(--accent-2));
    -webkit-background-clip: text; background-clip: text; color: transparent;
  }
  .sub{
    color: var(--muted); font-size: clamp(1rem, 2.4vw, 1.1rem);
    max-width: 30rem; line-height:1.6; margin: 0 auto 1.8rem;
  }
  .cta-row{ display:flex; gap: 0.9rem; flex-wrap:wrap; justify-content:center; }
  .cta-row .btn{ padding: 0.8rem 1.6rem; font-size: 0.98rem; }
  .tags{ display:flex; flex-wrap:wrap; gap:0.5rem; justify-content:center; }
  .tag{
    font-size: 0.78rem; font-weight:600; color: var(--muted);
    border: 1px solid var(--line); border-radius: 999px; padding: 0.35rem 0.8rem;
    background: rgba(255,255,255,0.03);
  }

  section{ width:100%; }
  .section-head{ text-align:center; margin-bottom: 1.8rem; }
  .section-head .eyebrow{
    color: var(--accent-2); font-weight:700; font-size:0.82rem;
    text-transform:uppercase; letter-spacing:0.08em; margin-bottom:0.5rem;
  }
  .section-head h2{ font-size: clamp(1.5rem, 4vw, 2rem); margin:0; }

  .story{
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line);
    border-radius: 24px;
    padding: clamp(1.6rem, 5vw, 2.4rem);
  }
  .story p{ color: var(--muted); line-height:1.75; font-size: 1rem; margin: 0 0 1rem; }
  .story p:last-child{ margin-bottom:0; }
  .story strong{ color: var(--text); }
  .story .highlight{ color: var(--accent-2); font-weight:600; }

  .feature-grid{
    display:grid; grid-template-columns: 1fr; gap: 1rem;
  }
  @media (min-width: 48rem){ .feature-grid{ grid-template-columns: repeat(3, 1fr); } }
  .feature-card{
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: 18px;
    padding: 1.4rem 1.5rem;
  }
  .feature-icon{
    width: 2.6rem; height:2.6rem; border-radius: 12px;
    display:flex; align-items:center; justify-content:center;
    font-size: 1.3rem; margin-bottom: 0.9rem;
  }
  .feature-icon.blue{ background: rgba(79,140,255,0.12); color: var(--accent); }
  .feature-icon.teal{ background: rgba(95,208,197,0.12); color: var(--accent-2); }
  .feature-icon.green{ background: rgba(61,220,151,0.12); color: var(--present); }
  .feature-icon.amber{ background: rgba(247,201,72,0.12); color: var(--onduty); }
  .feature-card h3{ font-size: 1.02rem; margin: 0 0 0.4rem; }
  .feature-card p{ color: var(--muted); font-size: 0.9rem; line-height:1.6; margin:0; }

  .feature-spotlight{
    display:flex; flex-direction:column; align-items:center; text-align:center; gap:0.9rem;
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line);
    border-radius: 20px;
    padding: clamp(1.8rem, 5vw, 2.4rem);
    margin-top: 1rem;
  }
  .feature-spotlight .feature-icon{ margin-bottom:0; }
  .feature-spotlight h3{ font-size: 1.08rem; margin: 0; }
  .feature-spotlight p{ color: var(--muted); line-height:1.7; font-size: 0.95rem; margin:0; max-width: 40rem; }

  .steps{
    display:grid; grid-template-columns: 1fr; gap: 1rem;
  }
  @media (min-width: 40rem){ .steps{ grid-template-columns: repeat(3, 1fr); } }
  .step{
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line);
    border-radius: 18px;
    padding: 1.5rem;
    text-align:center;
  }
  .step-num{
    width: 2.2rem; height:2.2rem; border-radius:50%; margin: 0 auto 0.8rem;
    background: linear-gradient(135deg, var(--accent), var(--accent-2));
    color:#0b0d12; font-weight:800; font-size:0.95rem;
    display:flex; align-items:center; justify-content:center;
  }
  .step h3{ font-size: 0.98rem; margin: 0 0 0.4rem; }
  .step p{ color: var(--muted); font-size: 0.88rem; line-height:1.6; margin:0; }

  .bottom-cta{
    width:100%; text-align:center;
    background: linear-gradient(135deg, rgba(79,140,255,0.14), rgba(95,208,197,0.14));
    border: 1px solid var(--line);
    border-radius: 24px;
    padding: clamp(2rem, 5vw, 2.6rem);
  }
  .bottom-cta h2{ font-size: clamp(1.4rem, 4vw, 1.8rem); margin: 0 0 0.6rem; }
  .bottom-cta p{ color: var(--muted); margin: 0 0 1.4rem; }

  footer{ text-align:center; color: var(--muted); font-size: 0.85rem; padding: 1rem 1rem 2rem; }
  footer strong{ color: var(--text); }
</style>
<div class="page">

  <div class="hero-card">
    <div class="kicker">Automatic Attendance Registrar</div>
    <h1>AAR</h1>
    <p class="sub">Tap a card, attendance is marked. Built to work online across your whole school, and just as well with no internet at all.</p>
    <div class="cta-row">
      <a class="btn btn-primary" href="/signup">Get Started</a>
      <a class="btn btn-ghost" href="/signin">Sign In</a>
    </div>
  </div>
  <div class="tags">
    <span class="tag">ESP32 + RC522</span>
    <span class="tag">WebSocket Sync</span>
    <span class="tag">Offline-Capable</span>
  </div>

  <section>
    <div class="section-head">
      <div class="eyebrow">Why AAR exists</div>
      <h2>Attendance shouldn't take five minutes</h2>
    </div>
    <div class="story">
      <p>
        Every school day starts the same way: a teacher reads names off a sheet, one by one,
        while a room full of students waits. It's slow, it's easy to fake, and the second the
        wifi drops, most "smart" attendance systems stop working entirely &mdash; right when
        the school needs them to just keep running.
      </p>
      <p>
        <strong>AAR was built to fix that.</strong> A student taps their card on a reader and
        that's it &mdash; attendance is marked in under a second. Every reader keeps a local
        copy of its roster, so it <span class="highlight">never depends on the internet to do
        its job</span>. The network is only there for syncing changes down and rosters back up
        &mdash; not for taking attendance itself.
      </p>
    </div>
  </section>

  <section>
    <div class="section-head">
      <div class="eyebrow">What you get</div>
      <h2>Built for how schools actually run</h2>
    </div>
    <div class="feature-grid">
      <div class="feature-card">
        <div class="feature-icon blue">&#9889;</div>
        <h3>Tap & Done</h3>
        <p>One RFID tap marks a student present instantly. No apps, no manual entry, no queue.</p>
      </div>
      <div class="feature-card">
        <div class="feature-icon green">&#128274;</div>
        <h3>Works Fully Offline</h3>
        <p>Once synced, a reader keeps taking attendance with zero internet. The network is a convenience, not a dependency.</p>
      </div>
      <div class="feature-card">
        <div class="feature-icon amber">&#128737;</div>
        <h3>Encrypted In Transit</h3>
        <p>Every connection between a reader and the backend runs over TLS, keeping student data private on the wire.</p>
      </div>
    </div>
    <div class="feature-spotlight">
      <div class="feature-icon teal">&#127760;</div>
      <div>
        <h3>Roams With Students</h3>
        <p>Schools regularly host events &mdash; camps, sports meets, special sessions &mdash; that temporarily move students into classrooms that aren't their own. Most attendance systems break the moment someone isn't where they're expected to be. When online, AAR doesn't care which reader a card is tapped on &mdash; attendance still gets marked correctly, no matter which room a student is actually in.</p>
      </div>
    </div>
  </section>

  <section>
    <div class="section-head">
      <div class="eyebrow">How it works</div>
      <h2>Three steps, under a second</h2>
    </div>
    <div class="steps">
      <div class="step">
        <div class="step-num">1</div>
        <h3>Tap</h3>
        <p>A student taps their card on any registered reader, anywhere in the school.</p>
      </div>
      <div class="step">
        <div class="step-num">2</div>
        <h3>Sync</h3>
        <p>The reader checks its local roster first, then syncs with the backend the moment it's online.</p>
      </div>
      <div class="step">
        <div class="step-num">3</div>
        <h3>Done</h3>
        <p>Attendance updates instantly and shows up on the dashboard &mdash; no waiting, no paperwork.</p>
      </div>
    </div>
  </section>

  <div class="bottom-cta">
    <h2>Ready to stop calling names?</h2>
    <p>Set up your first reader and see attendance mark itself.</p>
    <div class="cta-row">
      <a class="btn btn-primary" href="/signup">Get Started</a>
    </div>
  </div>

</div>
<footer>Developed by <strong>Ashutosh Peri</strong></footer>

<!-- Shown when the page is reached via a Supabase email verification link -->
<div class="overlay" id="authOverlay">
  <div class="modal">
    <div class="modal-icon" id="authIcon"></div>
    <h3 id="authTitle"></h3>
    <p id="authMessage"></p>
    <button type="button" class="btn btn-primary" id="authOkBtn">OK</button>
    <a class="btn btn-primary" id="authSigninBtn" href="/signin" style="display:none;">Go to Sign In</a>
  </div>
</div>
<script>
  (function(){
    var authOverlay = document.getElementById('authOverlay');
    document.getElementById('authOkBtn').addEventListener('click', function(){
      authOverlay.classList.remove('show');
    });

    function showAuthModal(kind, glyph, title, message, isSuccess){
      var icon = document.getElementById('authIcon');
      icon.className = 'modal-icon ' + kind;
      icon.innerHTML = glyph;
      document.getElementById('authTitle').textContent = title;
      document.getElementById('authMessage').textContent = message;
      document.getElementById('authOkBtn').style.display = isSuccess ? 'none' : 'inline-flex';
      document.getElementById('authSigninBtn').style.display = isSuccess ? 'inline-flex' : 'none';
      authOverlay.classList.add('show');
    }

    var rawHash = window.location.hash.startsWith('#') ? window.location.hash.slice(1) : '';
    if (!rawHash) return;
    var params = new URLSearchParams(rawHash);
    var errorCode = params.get('error_code');
    var errorDesc = params.get('error_description');
    var accessToken = params.get('access_token');

    if (errorCode || params.get('error')){
      var message = errorCode === 'otp_expired'
        ? "This link has expired or has already been used. Please request a new one and try again."
        : (errorDesc ? errorDesc.replace(/\\+/g, ' ') : "Something went wrong verifying your email.");
      showAuthModal('danger', '&times;', 'Link expired or invalid', message, false);
      history.replaceState(null, '', window.location.pathname);
    } else if (accessToken){
      showAuthModal('present', '&check;', 'Email verified!', "Your account is now active. You can sign in and get started.", true);
      history.replaceState(null, '', window.location.pathname);
    }
  })();
</script>
""",
    title="AAR — Automatic Attendance Registrar",
)


# ---------- Sign up page ----------
SIGNUP_PAGE = page_shell(
    NAV + f"""
<style>
  .name-row{{ display:flex; gap:0.8rem; }}
  .name-row > div{{ flex:1; }}
  .phone-row{{ display:flex; }}
  .phone-prefix{{
    display:flex; align-items:center; justify-content:center;
    padding: 0 0.9rem; border: 1px solid var(--line); border-right:none;
    border-radius: 10px 0 0 10px; background: rgba(255,255,255,0.06);
    color: var(--text); font-size:0.95rem; font-weight:600; flex:none;
  }}
  .phone-row input{{ border-radius: 0 10px 10px 0; }}
  .select-wrap{{ position:relative; }}
  .select-wrap::after{{
    content:""; position:absolute; right:1.1rem; top:50%;
    width:0; height:0; pointer-events:none;
    border-left:5px solid transparent; border-right:5px solid transparent;
    border-top:6px solid var(--accent-2);
    transition: transform 0.18s ease;
    transform: translateY(-50%);
  }}
  .select-wrap:focus-within::after{{ transform: translateY(-50%) rotate(180deg); }}
  .form-card select{{
    appearance:none; -webkit-appearance:none; -moz-appearance:none;
    cursor:pointer; padding-right:2.6rem;
  }}
  .form-card select:invalid{{ color: var(--muted); }}
  .form-card select:valid{{ color: var(--text); }}
  .form-card select option{{ background: var(--card); color: var(--text); padding: 0.5rem; }}
  .form-card select option[value=""]{{ color: var(--muted); }}
</style>
<div class="form-wrap">
  <div class="form-card">
    <h1>Create your AAR account</h1>
    <p class="form-sub">Tell us about your institute to get started.</p>
    <form id="signupForm">
      <div class="name-row">
        <div>
          <div class="field-label">First Name</div>
          <input type="text" name="first_name" required placeholder="e.g. John">
        </div>
        <div>
          <div class="field-label">Last Name</div>
          <input type="text" name="last_name" required placeholder="e.g. Doe">
        </div>
      </div>

      <div class="field-label">School Name</div>
      <input type="text" name="institute_name" required placeholder="e.g. School Name">

      <div class="field-label">Institute Branch Name <span class="optional">(optional)</span></div>
      <input type="text" name="institute_branch_name" placeholder="e.g. Sector 5">

      <div class="field-label">Email</div>
      <input type="email" name="email" required placeholder="you@institute.edu">

      <div class="name-row">
        <div>
          <div class="field-label">Password</div>
          <input type="password" name="password" required minlength="8" placeholder="At least 8 characters">
        </div>
        <div>
          <div class="field-label">Confirm Password</div>
          <input type="password" name="confirm_password" required minlength="8" placeholder="Re-enter password">
        </div>
      </div>

      <div class="field-label">Mobile Number</div>
      <div class="phone-row">
        <span class="phone-prefix">+91</span>
        <input type="tel" name="mobile_number" required placeholder="90000 00000"
               maxlength="10" pattern="[0-9]{{10}}" inputmode="numeric"
               title="Enter a 10-digit mobile number">
      </div>

      <div class="field-label">Number of Students</div>
      <div class="select-wrap">
        <select name="student_count_range" required>
          <option value="" disabled selected>Select a range</option>
          <option value="<250">Fewer than 250</option>
          <option value="250-500">250 &ndash; 500</option>
          <option value="500-1000">500 &ndash; 1000</option>
          <option value="1000-3000">1000 &ndash; 3000</option>
          <option value="3000+">3000+</option>
        </select>
      </div>

      <button type="submit" class="btn btn-primary form-submit">Create Account</button>
      <div class="form-note" id="signupNote"></div>
    </form>
    <p class="form-switch">Already have an account? <a href="/signin">Sign In</a></p>
  </div>
</div>
<script>
  var phoneInput = document.querySelector('input[name="mobile_number"]');
  phoneInput.addEventListener('input', function(){{
    this.value = this.value.replace(/\\D/g, '').slice(0, 10);
  }});

  document.getElementById('signupForm').addEventListener('submit', async function(e){{
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('signupNote');
    var btn = form.querySelector('button[type="submit"]');
    var data = Object.fromEntries(new FormData(form));

    if (data.password !== data.confirm_password){{
      note.style.color = 'var(--absent)';
      note.textContent = "Passwords don't match.";
      return;
    }}

    btn.disabled = true;
    note.style.color = 'var(--muted)';
    note.textContent = "Creating your account...";

    try{{
      var res = await fetch('/api/signup', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: JSON.stringify(data)
      }});
      var result = await res.json();
      if (result.ok){{
        note.style.color = 'var(--present)';
        note.textContent = "Account created! Check your email to verify, then sign in.";
        form.reset();
      }} else {{
        note.style.color = 'var(--absent)';
        note.textContent = result.error || "Something went wrong. Please try again.";
      }}
    }} catch (err){{
      note.style.color = 'var(--absent)';
      note.textContent = "Could not reach the server. Please try again.";
    }} finally {{
      btn.disabled = false;
    }}
  }});
</script>
""",
    title="Sign Up — AAR",
)


# ---------- Sign in page ----------
SIGNIN_PAGE = page_shell(
    NAV + """
<div class="form-wrap">
  <div class="form-card">
    <h1>Sign in to AAR</h1>
    <p class="form-sub">Welcome back.</p>
    <form id="signinForm">
      <div class="field-label">Email</div>
      <input type="email" name="email" required placeholder="you@institute.edu">

      <div class="field-label">Password</div>
      <input type="password" name="password" required placeholder="Your password">

      <div class="form-links">
        <button type="button" id="forgotBtn">Forgot Password?</button>
      </div>

      <button type="submit" class="btn btn-primary form-submit">Sign In</button>
      <div class="form-note" id="signinNote"></div>
    </form>
    <p class="form-switch">Don't have an account? <a href="/signup">Sign Up</a></p>
  </div>
</div>

<!-- Error modal: wrong credentials / not verified -->
<div class="overlay" id="errorOverlay">
  <div class="modal">
    <div class="modal-icon" id="errorIcon"></div>
    <h3 id="errorTitle"></h3>
    <p id="errorMessage"></p>
    <button type="button" class="btn btn-primary" id="errorOkBtn">OK</button>
  </div>
</div>

<!-- Forgot password modal -->
<div class="overlay" id="forgotOverlay">
  <div class="modal">
    <div class="modal-icon warn">&#9993;</div>
    <h3>Reset your password</h3>
    <p>Enter your account email and we'll send you a link to reset your password.</p>
    <form id="forgotForm">
      <input type="email" name="email" required placeholder="you@institute.edu">
      <button type="submit" class="btn btn-primary">Send Reset Link</button>
      <div class="modal-note" id="forgotNote"></div>
    </form>
  </div>
</div>

<script>
  function showOverlay(el){ el.classList.add('show'); }
  function hideOverlay(el){ el.classList.remove('show'); }

  document.querySelectorAll('.overlay').forEach(function(ov){
    ov.addEventListener('click', function(e){ if (e.target === ov) hideOverlay(ov); });
  });

  var errorOverlay = document.getElementById('errorOverlay');
  var errorIcon = document.getElementById('errorIcon');
  var errorTitle = document.getElementById('errorTitle');
  var errorMessage = document.getElementById('errorMessage');
  document.getElementById('errorOkBtn').addEventListener('click', function(){ hideOverlay(errorOverlay); });

  function showErrorModal(kind, title, message){
    errorIcon.className = 'modal-icon ' + kind;
    errorIcon.innerHTML = kind === 'danger' ? '&times;' : '!';
    errorTitle.textContent = title;
    errorMessage.textContent = message;
    showOverlay(errorOverlay);
  }

  document.getElementById('signinForm').addEventListener('submit', async function(e){
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('signinNote');
    var btn = form.querySelector('button[type="submit"]');
    var data = Object.fromEntries(new FormData(form));

    btn.disabled = true;
    note.textContent = '';

    try{
      var res = await fetch('/api/signin', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(data)
      });
      var result = await res.json();

      if (result.ok){
        localStorage.setItem('aar_access_token', result.access_token);
        localStorage.setItem('aar_refresh_token', result.refresh_token);
        window.location.href = '/dashboard';
        return;
      }

      if (result.reason === 'not_verified'){
        showErrorModal('warn', 'Email not yet verified',
          "Your account exists but hasn't been verified yet. Check your inbox for the verification link.");
      } else {
        showErrorModal('danger', 'Email ID or Password is wrong',
          "Double-check your email and password and try again.");
      }
    } catch (err){
      showErrorModal('danger', "Couldn't sign in",
        "Could not reach the server. Please try again.");
    } finally {
      btn.disabled = false;
    }
  });

  var forgotOverlay = document.getElementById('forgotOverlay');
  document.getElementById('forgotBtn').addEventListener('click', function(){
    document.getElementById('forgotNote').textContent = '';
    showOverlay(forgotOverlay);
  });

  document.getElementById('forgotForm').addEventListener('submit', async function(e){
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('forgotNote');
    var btn = form.querySelector('button[type="submit"]');
    var email = form.email.value.trim();
    if (!email) return;

    btn.disabled = true;
    note.textContent = '';
    try{
      await fetch('/api/forgot-password', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ email: email })
      });
    } catch (err){ /* still show the generic confirmation below */ }
    note.style.color = 'var(--present)';
    note.textContent = "If that email is registered, a reset link has been sent.";
    btn.disabled = false;
  });
</script>
""",
    title="Sign In — AAR",
)


# ---------- Dashboard (gated client-side by the stored Supabase token) ----------
DASHBOARD_PAGE = page_shell(
    NAV + """
<style>
  body{ overflow-x:hidden; }
  .dash-layout{ display:flex; align-items:flex-start; min-height: calc(100vh - 5rem); }
  .sidebar{
    width: 18rem; flex:none; align-self:stretch;
    background: var(--card); border-right: 1px solid var(--line);
    padding: 1.5rem 1rem; display:flex; flex-direction:column; gap:0.4rem;
    position:sticky; top:0;
  }
  .sidebar button{
    display:flex; align-items:center; gap:0.7rem;
    width:100%; text-align:left; background:none; border:none;
    color: var(--muted); font-size:0.92rem; font-weight:600;
    padding: 0.7rem 0.9rem; border-radius:10px; cursor:pointer;
    white-space:nowrap;
    transition: background 0.15s ease, color 0.15s ease;
  }
  .sidebar button:hover{ background: rgba(255,255,255,0.04); color: var(--text); }
  .sidebar button.active{
    background: linear-gradient(135deg, rgba(79,140,255,0.16), rgba(95,208,197,0.16));
    color: var(--accent-2); border: 1px solid rgba(79,140,255,0.3);
  }
  .sidebar .nav-icon{
    font-size:1.15rem; line-height:1; flex:none;
    display:inline-flex; align-items:center; justify-content:center;
  }
  .dash-content{ flex:1; min-width:0; padding: clamp(1.5rem, 4vw, 2.5rem); }
  .view{ display:none; }
  .view.active{ display:block; }
  .view h1{ font-size: clamp(1.4rem, 3.5vw, 1.7rem); margin: 0 0 1.4rem; }

  .stat-grid{ display:grid; grid-template-columns: 1fr; gap:1rem; }
  @media (min-width: 40rem){ .stat-grid{ grid-template-columns: repeat(3, 1fr); } }
  .stat-card{
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border:1px solid var(--line); border-radius:18px; padding:1.4rem 1.5rem;
  }
  .stat-card .stat-label{ font-size:0.76rem; color:var(--muted); text-transform:uppercase; letter-spacing:0.05em; margin-bottom:0.6rem; }
  .stat-card .stat-value{ font-size:2.1rem; font-weight:800; }
  .stat-card.blue .stat-value{ color: var(--accent); }
  .stat-card.green .stat-value{ color: var(--present); }
  .stat-card.teal .stat-value{ color: var(--accent-2); }

  .toolbar{ display:flex; justify-content:space-between; flex-wrap:wrap; gap:0.8rem; margin-bottom:1.4rem; }
  .toolbar-group{ display:flex; gap:0.6rem; flex-wrap:wrap; }
  .toolbar .btn{ font-size:0.88rem; padding:0.6rem 1rem; }

  .table-card{
    background: var(--card); border:1px solid var(--line); border-radius:18px;
    overflow:hidden;
  }
  .table-scroll{ overflow-x:auto; }
  .data-table{ width:100%; border-collapse:collapse; min-width:36rem; }
  .data-table th{
    text-align:left; font-size:0.72rem; text-transform:uppercase; letter-spacing:0.05em;
    color:var(--muted); padding:0.8rem 1rem; border-bottom:1px solid var(--line);
    white-space:nowrap;
  }
  .data-table td{ padding:0.8rem 1rem; border-bottom:1px solid var(--line); font-size:0.9rem; white-space:nowrap; }
  .data-table tr:last-child td{ border-bottom:none; }
  .data-table tr.student-row{ transition: background 0.4s ease, box-shadow 0.4s ease; }
  .data-table tr.student-row.highlight{ background: rgba(79,140,255,0.16); box-shadow: inset 0 0 0 1px rgba(79,140,255,0.45); }
  .aaruid-cell{ color:var(--muted); font-size:0.8rem; }
  .empty-row td{ text-align:center; color:var(--muted); padding:2.5rem 1rem; }

  .import-drop{
    border:2px dashed var(--line); border-radius:16px; padding:2.6rem 1.5rem;
    text-align:center; color:var(--muted); cursor:pointer;
    transition: border-color 0.2s ease, background 0.2s ease;
  }
  .import-drop.dragover{ border-color: var(--accent); background: rgba(79,140,255,0.07); }
  .import-drop .import-icon{ font-size:2rem; margin-bottom:0.8rem; color: var(--accent-2); }
  .import-hint{ font-size:0.78rem; color:var(--muted); margin-top:1rem; text-align:center; }

  .import-step{ display:none; }
  .import-step.active{ display:block; }
  .import-spinner{
    width:2.4rem; height:2.4rem; border-radius:50%; margin: 0 auto 1.2rem;
    border: 3px solid var(--line); border-top-color: var(--accent);
    animation: spin 0.8s linear infinite;
  }
  @keyframes spin{ to{ transform: rotate(360deg); } }
  .import-summary-row{ display:flex; justify-content:space-between; padding:0.5rem 0; font-size:0.9rem; border-bottom:1px solid var(--line); }
  .import-summary-row:last-child{ border-bottom:none; }
  .import-summary-row span:last-child{ font-weight:700; }
  .dup-choice{ display:flex; gap:0.6rem; margin-top:1rem; }
  .dup-choice button{
    flex:1; padding:0.7rem 0.5rem; border-radius:10px; border:1px solid var(--line);
    background: rgba(255,255,255,0.04); color:var(--text); font-size:0.85rem; font-weight:600; cursor:pointer;
  }
  .dup-choice button.selected{ border-color: var(--accent); color: var(--accent-2); background: rgba(79,140,255,0.1); }

  .template-preview{ width:100%; border-collapse:collapse; margin-top:1rem; }
  .template-preview th, .template-preview td{
    border:1px solid var(--line); padding:0.55rem 0.7rem; font-size:0.82rem; text-align:left;
  }
  .template-preview th{ background: rgba(255,255,255,0.04); color: var(--accent-2); }

  .modal.wide{ max-width:30rem; text-align:left; }
  .modal.wide h3{ text-align:center; }

  .review-card{
    max-width: 36rem; margin: 0 auto; text-align:center;
    background: linear-gradient(160deg, var(--card), var(--card-2));
    border: 1px solid var(--line); border-radius: 20px;
    padding: 2.4rem;
  }
  .review-image-wrap{
    position:relative; width:100%; aspect-ratio: 4 / 3;
    background: rgba(255,255,255,0.03); border: 1px solid var(--line); border-radius: 14px;
    overflow:hidden; display:flex; align-items:center; justify-content:center;
  }
  .review-image-wrap img{ width:100%; height:100%; object-fit:contain; }
  .review-placeholder{ color: var(--muted); font-size:0.85rem; }
  .review-name{ font-size:1.6rem; font-weight:800; letter-spacing:0.02em; margin:1.4rem 0 1rem; }
  .review-nav{ display:flex; align-items:center; justify-content:center; gap:1.4rem; }
  .review-arrow{ font-size:1.4rem; padding:0.6rem 1.3rem; }
  .review-position{ color: var(--muted); font-size:0.9rem; min-width:3.5rem; }
</style>

<div class="dash-layout">
  <div class="sidebar">
    <button type="button" class="nav-btn active" data-view="home">
      <span class="nav-icon">&#127968;</span> Home
    </button>
    <button type="button" class="nav-btn" data-view="students">
      <span class="nav-icon">&#128101;</span> Student Management
    </button>
    <button type="button" class="nav-btn" data-view="devices">
      <span class="nav-icon">&#128225;</span> Device Manager
    </button>
    <button type="button" class="nav-btn" data-view="review">
      <span class="nav-icon">&#128248;</span> Review
    </button>
  </div>

  <div class="dash-content">

    <div class="view active" id="homeView">
      <h1>Home</h1>
      <div class="stat-grid">
        <div class="stat-card blue">
          <div class="stat-label">Devices Online</div>
          <div class="stat-value" id="statDevices">&mdash;</div>
        </div>
        <div class="stat-card green">
          <div class="stat-label">Students Registered</div>
          <div class="stat-value" id="statStudents">&mdash;</div>
        </div>
        <div class="stat-card teal">
          <div class="stat-label">Classes</div>
          <div class="stat-value" id="statClasses">&mdash;</div>
        </div>
      </div>
    </div>

    <div class="view" id="studentsView">
      <h1>Student Management</h1>
      <div class="toolbar">
        <div class="toolbar-group">
          <button type="button" class="btn btn-primary" id="openCreateBtn">&#43; Create</button>
          <button type="button" class="btn btn-ghost" id="openSearchBtn">&#128269; Search</button>
        </div>
        <div class="toolbar-group">
          <button type="button" class="btn btn-ghost" id="openTemplateBtn">Template</button>
          <button type="button" class="btn btn-ghost" id="openImportBtn">Import</button>
        </div>
      </div>

      <div class="table-card">
        <div class="table-scroll">
          <table class="data-table">
            <thead>
              <tr>
                <th>Name</th><th>Class</th><th>Section</th><th>RFID</th><th>AAR UID</th>
              </tr>
            </thead>
            <tbody id="studentsTableBody">
              <tr class="empty-row"><td colspan="5">Loading students&hellip;</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>

    <div class="view" id="devicesView">
      <h1>Device Manager</h1>
      <div class="toolbar">
        <div class="toolbar-group">
          <button type="button" class="btn btn-primary" id="openRegisterDeviceBtn">&#43; Register Device</button>
        </div>
      </div>

      <div class="table-card">
        <div class="table-scroll">
          <table class="data-table">
            <thead>
              <tr>
                <th>Name</th><th>Class</th><th>Section</th><th>MAC Address</th><th>AAR UID</th><th>Actions</th>
              </tr>
            </thead>
            <tbody id="devicesTableBody">
              <tr class="empty-row"><td colspan="6">Loading devices&hellip;</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </div>

    <div class="view" id="reviewView">
      <h1>Review</h1>
      <div class="review-card">
        <div class="review-image-wrap">
          <img id="reviewImage" alt="" style="display:none;">
          <div class="review-placeholder" id="reviewPlaceholder">Loading&hellip;</div>
        </div>
        <div class="review-name" id="reviewName">&mdash;</div>
        <div class="review-nav">
          <button type="button" class="btn btn-ghost review-arrow" id="reviewPrevBtn">&larr;</button>
          <span class="review-position" id="reviewPosition"></span>
          <button type="button" class="btn btn-ghost review-arrow" id="reviewNextBtn">&rarr;</button>
        </div>
      </div>
    </div>

  </div>
</div>

<!-- Register device -->
<div class="overlay" id="registerDeviceOverlay">
  <div class="modal wide">
    <h3>Register Device</h3>
    <form id="registerDeviceForm">
      <div class="field-label">Device MAC</div>
      <input type="text" name="mac_address" required placeholder="e.g. AC:23:F0:12:34:56">
      <div class="field-label">Device Name</div>
      <input type="text" name="name" required placeholder="e.g. Reader - Main Gate">
      <div class="field-label">Device Class</div>
      <input type="text" name="class" required placeholder="e.g. 8">
      <div class="field-label">Device Section</div>
      <input type="text" name="section" required placeholder="e.g. A">
      <div class="modal-note" id="registerDeviceNote"></div>
      <div class="dup-choice" style="margin-top:1.4rem;">
        <button type="button" class="btn btn-ghost" id="cancelRegisterDeviceBtn" style="flex:1; text-align:center;">Cancel</button>
        <button type="submit" class="btn btn-primary" style="flex:1;">Register</button>
      </div>
    </form>
  </div>
</div>

<!-- Edit device -->
<div class="overlay" id="editDeviceOverlay">
  <div class="modal wide">
    <h3>Edit Device</h3>
    <form id="editDeviceForm">
      <div class="field-label">Device MAC <span class="optional">(cannot be changed)</span></div>
      <input type="text" id="editDeviceMac" disabled style="color:var(--muted); cursor:not-allowed;">
      <div class="field-label">Device Name</div>
      <input type="text" name="name" required placeholder="e.g. Reader - Main Gate">
      <div class="field-label">Device Class</div>
      <input type="text" name="class" required placeholder="e.g. 8">
      <div class="field-label">Device Section</div>
      <input type="text" name="section" required placeholder="e.g. A">
      <div class="modal-note" id="editDeviceNote"></div>
      <div class="dup-choice" style="margin-top:1.4rem;">
        <button type="button" class="btn btn-ghost" id="cancelEditDeviceBtn" style="flex:1; text-align:center;">Cancel</button>
        <button type="submit" class="btn btn-primary" style="flex:1;">Save Changes</button>
      </div>
    </form>
  </div>
</div>

<!-- Delete device confirmation -->
<div class="overlay" id="deleteDeviceOverlay">
  <div class="modal">
    <div class="modal-icon danger">&#33;</div>
    <h3>Delete this device?</h3>
    <p id="deleteDeviceMessage">This cannot be undone.</p>
    <div class="dup-choice">
      <button type="button" class="btn btn-ghost" id="cancelDeleteDeviceBtn" style="flex:1;">Cancel</button>
      <button type="button" class="btn btn-primary" id="confirmDeleteDeviceBtn" style="flex:1; background:var(--absent);">Delete</button>
    </div>
    <div class="modal-note" id="deleteDeviceNote"></div>
  </div>
</div>

<!-- Create student -->
<div class="overlay" id="createOverlay">
  <div class="modal wide">
    <h3>Create Student</h3>
    <form id="createForm">
      <div class="field-label">Name</div>
      <input type="text" name="name" required placeholder="e.g. John Doe">
      <div class="field-label">RFID</div>
      <input type="text" name="rfid" required placeholder="e.g. AA:BB:CC:DD">
      <div class="field-label">Class</div>
      <input type="text" name="class" required placeholder="e.g. 8">
      <div class="field-label">Section</div>
      <input type="text" name="section" required placeholder="e.g. A">
      <div class="modal-note" id="createNote"></div>
      <div class="dup-choice" style="margin-top:1.4rem;">
        <button type="button" class="btn btn-ghost" id="cancelCreateBtn" style="flex:1; text-align:center;">Cancel</button>
        <button type="submit" class="btn btn-primary" style="flex:1;">Create</button>
      </div>
    </form>
  </div>
</div>

<!-- Search student -->
<div class="overlay" id="searchOverlay">
  <div class="modal">
    <div class="modal-icon warn">&#128269;</div>
    <h3>Search Students</h3>
    <p>Enter a student's name, RFID, class, or section.</p>
    <form id="searchForm">
      <input type="text" name="query" required placeholder="e.g. John Doe or AA:BB:CC:DD">
      <button type="submit" class="btn btn-primary">Search</button>
      <div class="modal-note" id="searchNote"></div>
    </form>
  </div>
</div>

<!-- Template preview -->
<div class="overlay" id="templateOverlay">
  <div class="modal wide">
    <h3>Import Template Format</h3>
    <p style="color:var(--muted); font-size:0.88rem; margin:0 0 0.5rem;">
      Your spreadsheet's first row must have these exact column headers (any order):
    </p>
    <table class="template-preview">
      <thead><tr><th>Name</th><th>RFID</th><th>Class</th><th>Section</th></tr></thead>
      <tbody><tr><td>John Doe</td><td>AA:BB:CC:DD</td><td>8</td><td>A</td></tr></tbody>
    </table>
    <div class="dup-choice" style="margin-top:1.4rem;">
      <a class="btn btn-ghost" href="/api/students/template" style="flex:1; text-align:center;">Download Template</a>
      <button type="button" class="btn btn-primary" id="closeTemplateBtn" style="flex:1;">Close</button>
    </div>
  </div>
</div>

<!-- Import students -->
<div class="overlay" id="importOverlay">
  <div class="modal wide">
    <h3>Import Students</h3>

    <div class="import-step active" id="importStepUpload">
      <div class="import-drop" id="importDrop">
        <div class="import-icon">&#8593;</div>
        <div>Drag and drop an Excel file here</div>
        <div class="import-hint">or</div>
        <div style="margin-top:0.8rem;">
          <button type="button" class="btn btn-primary" id="importBrowseBtn">Choose File</button>
        </div>
      </div>
      <input type="file" id="importFileInput" accept=".xlsx" style="display:none;">
      <p class="import-hint">Accepts .xlsx files matching the import template.</p>
      <div class="modal-note" id="importUploadNote"></div>
    </div>

    <div class="import-step" id="importStepAnalysing">
      <div class="import-spinner"></div>
      <p style="text-align:center;">Analysing&hellip;</p>
    </div>

    <div class="import-step" id="importStepConfirm">
      <div class="import-summary-row"><span>Students found</span><span id="confirmTotal">0</span></div>
      <div class="import-summary-row"><span>New students</span><span id="confirmNew">0</span></div>
      <div class="import-summary-row" id="confirmDupRow" style="display:none;">
        <span>Duplicate RFIDs found</span><span id="confirmDup">0</span>
      </div>
      <div id="dupChoiceWrap" style="display:none;">
        <p style="color:var(--muted); font-size:0.85rem; margin:1rem 0 0;">What should happen to duplicates?</p>
        <div class="dup-choice">
          <button type="button" class="dup-option selected" data-strategy="skip">Skip Adding</button>
          <button type="button" class="dup-option" data-strategy="replace">Replace</button>
        </div>
      </div>
      <div class="dup-choice" style="margin-top:1.4rem;">
        <button type="button" class="btn btn-ghost" id="cancelImportBtn" style="flex:1;">Cancel</button>
        <button type="button" class="btn btn-primary" id="confirmImportBtn" style="flex:1;">Confirm Import</button>
      </div>
      <div class="modal-note" id="importConfirmNote"></div>
    </div>

    <div class="import-step" id="importStepDone">
      <div class="modal-icon present" style="margin:0 auto 1rem;">&#10003;</div>
      <p style="text-align:center;" id="importDoneMessage"></p>
      <button type="button" class="btn btn-primary" id="closeImportBtn" style="width:100%;">Done</button>
    </div>
  </div>
</div>

<script>
(function(){
  var token = localStorage.getItem('aar_access_token');
  var refreshToken = localStorage.getItem('aar_refresh_token');
  if (!token){
    window.location.href = '/signin';
    return;
  }

  function signOutAndRedirect(){
    localStorage.removeItem('aar_access_token');
    localStorage.removeItem('aar_refresh_token');
    window.location.href = '/signin';
  }

  // Access tokens are short-lived. If one has expired, use the refresh token
  // to get a new pair and retry once, instead of silently 401-ing everything
  // and leaving the page looking "logged in" but broken. Concurrent requests
  // share one in-flight refresh instead of each triggering their own.
  var refreshPromise = null;
  function refreshSession(){
    if (!refreshPromise){
      refreshPromise = fetch('/api/refresh', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ refresh_token: refreshToken })
      }).then(function(r){ return r.json(); }).then(function(data){
        refreshPromise = null;
        if (!data.ok){
          signOutAndRedirect();
          throw new Error('session expired');
        }
        token = data.access_token;
        refreshToken = data.refresh_token;
        localStorage.setItem('aar_access_token', token);
        localStorage.setItem('aar_refresh_token', refreshToken);
        return token;
      }).catch(function(err){
        refreshPromise = null;
        throw err;
      });
    }
    return refreshPromise;
  }

  function authFetch(url, options){
    options = options || {};
    options.headers = Object.assign({}, options.headers, { 'Authorization': 'Bearer ' + token });
    return fetch(url, options).then(function(res){
      if (res.status !== 401) return res;
      return refreshSession().then(function(newToken){
        options.headers['Authorization'] = 'Bearer ' + newToken;
        return fetch(url, options);
      });
    });
  }

  // ---------- Sidebar tab switching ----------
  // The active tab is tracked in the URL hash so a page refresh (or a
  // shared/bookmarked link) lands back on the same tab instead of always
  // resetting to Home.
  var navButtons = document.querySelectorAll('.nav-btn');
  var views = document.querySelectorAll('.view');
  var validViews = ['home', 'students', 'devices', 'review'];

  function switchToView(viewName, updateHash){
    if (validViews.indexOf(viewName) === -1) viewName = 'home';
    navButtons.forEach(function(b){ b.classList.toggle('active', b.getAttribute('data-view') === viewName); });
    views.forEach(function(v){ v.classList.remove('active'); });
    document.getElementById(viewName + 'View').classList.add('active');
    if (updateHash !== false) window.location.hash = viewName;
    return viewName;
  }

  function activateView(viewName, updateHash){
    viewName = switchToView(viewName, updateHash);
    if (viewName === 'home') loadStats();
    if (viewName === 'students') loadStudents();
    if (viewName === 'devices') loadDevices();
    if (viewName === 'review') showReviewImage();
  }

  navButtons.forEach(function(btn){
    btn.addEventListener('click', function(){ activateView(btn.getAttribute('data-view')); });
  });

  window.addEventListener('hashchange', function(){
    activateView(window.location.hash.replace('#', ''), false);
  });

  activateView(window.location.hash.replace('#', ''), false);

  // ---------- Home stats ----------
  function loadStats(){
    authFetch('/api/dashboard/stats').then(function(r){ return r.json(); }).then(function(data){
      if (!data.ok){
        ['statDevices', 'statStudents', 'statClasses'].forEach(function(id){
          document.getElementById(id).textContent = '!';
        });
        return;
      }
      document.getElementById('statDevices').textContent = data.devices_online;
      document.getElementById('statStudents').textContent = data.students_registered;
      document.getElementById('statClasses').textContent = data.classes_count;
    }).catch(function(){
      ['statDevices', 'statStudents', 'statClasses'].forEach(function(id){
        document.getElementById(id).textContent = '!';
      });
    });
  }

  // ---------- Students table ----------
  var studentsCache = [];
  function renderStudents(students){
    var tbody = document.getElementById('studentsTableBody');
    tbody.innerHTML = '';
    if (students.length === 0){
      tbody.innerHTML = '<tr class="empty-row"><td colspan="5">No students yet. Click Create or Import to add some.</td></tr>';
      return;
    }
    students.forEach(function(s){
      var tr = document.createElement('tr');
      tr.className = 'student-row';
      tr.id = 'student-row-' + s.aaruid;
      tr.setAttribute('data-search', (s.name + ' ' + s.rfid + ' ' + s.class + ' ' + s.section).toLowerCase());
      tr.innerHTML =
        '<td>' + s.name + '</td>' +
        '<td>' + s.class + '</td>' +
        '<td>' + s.section + '</td>' +
        '<td>' + s.rfid + '</td>' +
        '<td class="aaruid-cell">' + s.aaruid + '</td>';
      tbody.appendChild(tr);
    });
  }

  function loadStudents(){
    var tbody = document.getElementById('studentsTableBody');
    authFetch('/api/students').then(function(r){ return r.json(); }).then(function(data){
      if (!data.ok){
        tbody.innerHTML = '<tr class="empty-row"><td colspan="5">Could not load students. Try refreshing the page.</td></tr>';
        return;
      }
      studentsCache = data.students;
      renderStudents(studentsCache);
    }).catch(function(){
      tbody.innerHTML = '<tr class="empty-row"><td colspan="5">Could not reach the server.</td></tr>';
    });
  }

  // ---------- Devices table ----------
  var devicesCache = [];
  function renderDevices(devices){
    var tbody = document.getElementById('devicesTableBody');
    tbody.innerHTML = '';
    if (devices.length === 0){
      tbody.innerHTML = '<tr class="empty-row"><td colspan="6">No devices yet. Click Register Device to add one.</td></tr>';
      return;
    }
    devices.forEach(function(d){
      var tr = document.createElement('tr');
      tr.className = 'student-row';
      tr.innerHTML =
        '<td>' + d.name + '</td>' +
        '<td>' + d.class + '</td>' +
        '<td>' + d.section + '</td>' +
        '<td>' + d.mac_address + '</td>' +
        '<td class="aaruid-cell">' + d.aaruid + '</td>' +
        '<td>' +
          '<button type="button" class="btn btn-ghost row-action" data-action="edit-device" data-aaruid="' + d.aaruid + '" style="padding:0.35rem 0.7rem; font-size:0.8rem;">Edit</button> ' +
          '<button type="button" class="btn btn-ghost row-action" data-action="delete-device" data-aaruid="' + d.aaruid + '" style="padding:0.35rem 0.7rem; font-size:0.8rem; color:var(--absent);">Delete</button>' +
        '</td>';
      tbody.appendChild(tr);
    });
  }

  function loadDevices(){
    var tbody = document.getElementById('devicesTableBody');
    authFetch('/api/devices').then(function(r){ return r.json(); }).then(function(data){
      if (!data.ok){
        tbody.innerHTML = '<tr class="empty-row"><td colspan="6">Could not load devices. Try refreshing the page.</td></tr>';
        return;
      }
      devicesCache = data.devices;
      renderDevices(devicesCache);
    }).catch(function(){
      tbody.innerHTML = '<tr class="empty-row"><td colspan="6">Could not reach the server.</td></tr>';
    });
  }

  // ---------- Review gallery ----------
  // Images are cached via the Cache Storage API as soon as the dashboard
  // loads (not just when the Review tab is opened), so by the time someone
  // actually clicks Review, every image renders instantly from cache
  // instead of waiting on a fresh network fetch.
  var IMAGE_CACHE_NAME = 'aar-review-images-v1';
  var reviewImages = [];
  var reviewIndex = 0;
  var reviewImagesPromise = null;

  function getReviewImages(){
    if (!reviewImagesPromise){
      reviewImagesPromise = authFetch('/api/gallery').then(function(r){ return r.json(); }).then(function(data){
        return data.ok ? data.images : [];
      }).catch(function(){ return []; });
    }
    return reviewImagesPromise;
  }

  function cacheReviewImages(images){
    if (!window.caches) return;
    caches.open(IMAGE_CACHE_NAME).then(function(cache){
      images.forEach(function(item){
        cache.match(item.url).then(function(existing){
          if (existing) return;
          fetch(item.url).then(function(res){
            if (res.ok) cache.put(item.url, res.clone());
          }).catch(function(){});
        });
      });
    });
  }

  function renderReviewImage(url, imgEl){
    if (!window.caches){
      imgEl.src = url;
      return;
    }
    caches.open(IMAGE_CACHE_NAME).then(function(cache){ return cache.match(url); }).then(function(cached){
      if (cached){
        return cached.blob().then(function(blob){ imgEl.src = URL.createObjectURL(blob); });
      }
      imgEl.src = url;
      fetch(url).then(function(res){
        if (res.ok) caches.open(IMAGE_CACHE_NAME).then(function(cache){ cache.put(url, res.clone()); });
      }).catch(function(){});
    });
  }

  function showReviewImage(){
    var img = document.getElementById('reviewImage');
    var placeholder = document.getElementById('reviewPlaceholder');
    var nameEl = document.getElementById('reviewName');
    var posEl = document.getElementById('reviewPosition');

    getReviewImages().then(function(images){
      reviewImages = images;
      if (reviewImages.length === 0){
        img.style.display = 'none';
        placeholder.style.display = 'flex';
        placeholder.textContent = 'No review images found.';
        nameEl.textContent = '—';
        posEl.textContent = '';
        return;
      }
      if (reviewIndex >= reviewImages.length) reviewIndex = 0;
      var item = reviewImages[reviewIndex];
      placeholder.style.display = 'none';
      img.style.display = 'block';
      renderReviewImage(item.url, img);
      nameEl.textContent = item.name;
      posEl.textContent = (reviewIndex + 1) + ' / ' + reviewImages.length;
    });
  }

  document.getElementById('reviewPrevBtn').addEventListener('click', function(){
    if (reviewImages.length === 0) return;
    reviewIndex = (reviewIndex - 1 + reviewImages.length) % reviewImages.length;
    showReviewImage();
  });
  document.getElementById('reviewNextBtn').addEventListener('click', function(){
    if (reviewImages.length === 0) return;
    reviewIndex = (reviewIndex + 1) % reviewImages.length;
    showReviewImage();
  });

  // Warm the cache immediately on login/dashboard load, regardless of which
  // tab is currently active.
  getReviewImages().then(cacheReviewImages);

  // ---------- Overlay helpers ----------
  function showOverlay(el){ el.classList.add('show'); }
  function hideOverlay(el){ el.classList.remove('show'); }
  document.querySelectorAll('.overlay').forEach(function(ov){
    ov.addEventListener('click', function(e){ if (e.target === ov) hideOverlay(ov); });
  });

  // ---------- Create ----------
  var createOverlay = document.getElementById('createOverlay');
  document.getElementById('openCreateBtn').addEventListener('click', function(){
    document.getElementById('createForm').reset();
    document.getElementById('createNote').textContent = '';
    showOverlay(createOverlay);
  });
  document.getElementById('cancelCreateBtn').addEventListener('click', function(){ hideOverlay(createOverlay); });

  document.getElementById('createForm').addEventListener('submit', async function(e){
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('createNote');
    var data = Object.fromEntries(new FormData(form));
    var btn = form.querySelector('button[type="submit"]');
    btn.disabled = true;
    try{
      var res = await authFetch('/api/students', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(data)
      });
      var result = await res.json();
      if (result.ok){
        hideOverlay(createOverlay);
        loadStudents();
      } else {
        note.style.color = 'var(--absent)';
        note.textContent = result.error || 'Something went wrong.';
      }
    } catch (err){
      note.style.color = 'var(--absent)';
      note.textContent = 'Could not reach the server.';
    } finally {
      btn.disabled = false;
    }
  });

  // ---------- Register Device ----------
  var registerDeviceOverlay = document.getElementById('registerDeviceOverlay');
  document.getElementById('openRegisterDeviceBtn').addEventListener('click', function(){
    document.getElementById('registerDeviceForm').reset();
    document.getElementById('registerDeviceNote').textContent = '';
    showOverlay(registerDeviceOverlay);
  });
  document.getElementById('cancelRegisterDeviceBtn').addEventListener('click', function(){ hideOverlay(registerDeviceOverlay); });

  document.getElementById('registerDeviceForm').addEventListener('submit', async function(e){
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('registerDeviceNote');
    var data = Object.fromEntries(new FormData(form));
    var btn = form.querySelector('button[type="submit"]');
    btn.disabled = true;
    try{
      var res = await authFetch('/api/devices', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(data)
      });
      var result = await res.json();
      if (result.ok){
        hideOverlay(registerDeviceOverlay);
        loadDevices();
      } else {
        note.style.color = 'var(--absent)';
        note.textContent = result.error || 'Something went wrong.';
      }
    } catch (err){
      note.style.color = 'var(--absent)';
      note.textContent = 'Could not reach the server.';
    } finally {
      btn.disabled = false;
    }
  });

  // ---------- Edit / Delete device ----------
  // Edit and Delete buttons are added dynamically inside renderDevices(), so
  // clicks are handled via delegation on the table body rather than direct
  // listeners (which wouldn't exist yet for rows rendered after page load).
  var editDeviceOverlay = document.getElementById('editDeviceOverlay');
  var deleteDeviceOverlay = document.getElementById('deleteDeviceOverlay');
  var editingDeviceAaruid = null;
  var deletingDeviceAaruid = null;

  document.getElementById('devicesTableBody').addEventListener('click', function(e){
    var btn = e.target.closest('.row-action');
    if (!btn) return;
    var aaruid = btn.getAttribute('data-aaruid');
    var device = devicesCache.find(function(d){ return d.aaruid === aaruid; });
    if (!device) return;

    if (btn.getAttribute('data-action') === 'edit-device'){
      editingDeviceAaruid = aaruid;
      var form = document.getElementById('editDeviceForm');
      form.reset();
      document.getElementById('editDeviceMac').value = device.mac_address;
      form.name.value = device.name;
      form.class.value = device.class;
      form.section.value = device.section;
      document.getElementById('editDeviceNote').textContent = '';
      showOverlay(editDeviceOverlay);
    }

    if (btn.getAttribute('data-action') === 'delete-device'){
      deletingDeviceAaruid = aaruid;
      document.getElementById('deleteDeviceMessage').textContent =
        'Delete "' + device.name + '" (' + device.mac_address + ')? This cannot be undone.';
      document.getElementById('deleteDeviceNote').textContent = '';
      showOverlay(deleteDeviceOverlay);
    }
  });

  document.getElementById('cancelEditDeviceBtn').addEventListener('click', function(){ hideOverlay(editDeviceOverlay); });

  document.getElementById('editDeviceForm').addEventListener('submit', async function(e){
    e.preventDefault();
    var form = e.target;
    var note = document.getElementById('editDeviceNote');
    var data = Object.fromEntries(new FormData(form));
    var btn = form.querySelector('button[type="submit"]');
    btn.disabled = true;
    try{
      var res = await authFetch('/api/devices/' + encodeURIComponent(editingDeviceAaruid), {
        method: 'PATCH',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(data)
      });
      var result = await res.json();
      if (result.ok){
        hideOverlay(editDeviceOverlay);
        loadDevices();
      } else {
        note.style.color = 'var(--absent)';
        note.textContent = result.error || 'Something went wrong.';
      }
    } catch (err){
      note.style.color = 'var(--absent)';
      note.textContent = 'Could not reach the server.';
    } finally {
      btn.disabled = false;
    }
  });

  document.getElementById('cancelDeleteDeviceBtn').addEventListener('click', function(){ hideOverlay(deleteDeviceOverlay); });

  document.getElementById('confirmDeleteDeviceBtn').addEventListener('click', async function(){
    var note = document.getElementById('deleteDeviceNote');
    var btn = this;
    btn.disabled = true;
    try{
      var res = await authFetch('/api/devices/' + encodeURIComponent(deletingDeviceAaruid), { method: 'DELETE' });
      var result = await res.json();
      if (result.ok){
        hideOverlay(deleteDeviceOverlay);
        loadDevices();
      } else {
        note.style.color = 'var(--absent)';
        note.textContent = result.error || 'Something went wrong.';
      }
    } catch (err){
      note.style.color = 'var(--absent)';
      note.textContent = 'Could not reach the server.';
    } finally {
      btn.disabled = false;
    }
  });

  // ---------- Search ----------
  var searchOverlay = document.getElementById('searchOverlay');
  document.getElementById('openSearchBtn').addEventListener('click', function(){
    document.getElementById('searchForm').reset();
    document.getElementById('searchNote').textContent = '';
    showOverlay(searchOverlay);
  });

  document.getElementById('searchForm').addEventListener('submit', function(e){
    e.preventDefault();
    var query = e.target.query.value.trim().toLowerCase();
    var note = document.getElementById('searchNote');
    if (!query) return;

    var rows = document.querySelectorAll('#studentsTableBody tr.student-row');
    var match = null;
    rows.forEach(function(row){
      if (!match && row.getAttribute('data-search').indexOf(query) !== -1) match = row;
    });

    if (!match){
      note.style.color = 'var(--absent)';
      note.textContent = 'No matching student found.';
      return;
    }

    hideOverlay(searchOverlay);
    switchToView('students');

    match.scrollIntoView({ behavior: 'smooth', block: 'center' });
    match.classList.add('highlight');
    setTimeout(function(){ match.classList.remove('highlight'); }, 2600);
  });

  // ---------- Template ----------
  var templateOverlay = document.getElementById('templateOverlay');
  document.getElementById('openTemplateBtn').addEventListener('click', function(){ showOverlay(templateOverlay); });
  document.getElementById('closeTemplateBtn').addEventListener('click', function(){ hideOverlay(templateOverlay); });

  // ---------- Import ----------
  var importOverlay = document.getElementById('importOverlay');
  var importSteps = document.querySelectorAll('.import-step');
  function showImportStep(id){
    importSteps.forEach(function(s){ s.classList.remove('active'); });
    document.getElementById(id).classList.add('active');
  }

  var pendingImportToken = null;
  var dupStrategy = 'skip';

  function resetImportModal(){
    showImportStep('importStepUpload');
    document.getElementById('importUploadNote').textContent = '';
    document.getElementById('importFileInput').value = '';
    pendingImportToken = null;
    dupStrategy = 'skip';
    document.querySelectorAll('.dup-option').forEach(function(b){ b.classList.toggle('selected', b.getAttribute('data-strategy') === 'skip'); });
  }

  document.getElementById('openImportBtn').addEventListener('click', function(){
    resetImportModal();
    showOverlay(importOverlay);
  });
  document.getElementById('cancelImportBtn').addEventListener('click', function(){ hideOverlay(importOverlay); });
  document.getElementById('closeImportBtn').addEventListener('click', function(){
    hideOverlay(importOverlay);
    loadStudents();
    loadStats();
  });

  var importDrop = document.getElementById('importDrop');
  var importFileInput = document.getElementById('importFileInput');
  document.getElementById('importBrowseBtn').addEventListener('click', function(){ importFileInput.click(); });
  importFileInput.addEventListener('change', function(){
    if (importFileInput.files[0]) analyzeImportFile(importFileInput.files[0]);
  });

  importDrop.addEventListener('dragover', function(e){ e.preventDefault(); importDrop.classList.add('dragover'); });
  importDrop.addEventListener('dragleave', function(){ importDrop.classList.remove('dragover'); });
  importDrop.addEventListener('drop', function(e){
    e.preventDefault();
    importDrop.classList.remove('dragover');
    if (e.dataTransfer.files[0]) analyzeImportFile(e.dataTransfer.files[0]);
  });

  async function analyzeImportFile(file){
    showImportStep('importStepAnalysing');
    var formData = new FormData();
    formData.append('file', file);
    try{
      var res = await authFetch('/api/students/import/analyze', { method: 'POST', body: formData });
      var result = await res.json();
      if (!result.ok){
        showImportStep('importStepUpload');
        document.getElementById('importUploadNote').style.color = 'var(--absent)';
        document.getElementById('importUploadNote').textContent = result.error || 'Could not analyse the file.';
        return;
      }
      pendingImportToken = result.import_token;
      document.getElementById('confirmTotal').textContent = result.total;
      document.getElementById('confirmNew').textContent = result.new_count;
      var dupRow = document.getElementById('confirmDupRow');
      var dupWrap = document.getElementById('dupChoiceWrap');
      if (result.duplicate_count > 0){
        dupRow.style.display = 'flex';
        dupWrap.style.display = 'block';
        document.getElementById('confirmDup').textContent = result.duplicate_count;
      } else {
        dupRow.style.display = 'none';
        dupWrap.style.display = 'none';
      }
      document.getElementById('importConfirmNote').textContent = '';
      showImportStep('importStepConfirm');
    } catch (err){
      showImportStep('importStepUpload');
      document.getElementById('importUploadNote').style.color = 'var(--absent)';
      document.getElementById('importUploadNote').textContent = 'Could not reach the server.';
    }
  }

  document.querySelectorAll('.dup-option').forEach(function(btn){
    btn.addEventListener('click', function(){
      document.querySelectorAll('.dup-option').forEach(function(b){ b.classList.remove('selected'); });
      btn.classList.add('selected');
      dupStrategy = btn.getAttribute('data-strategy');
    });
  });

  document.getElementById('confirmImportBtn').addEventListener('click', async function(){
    var note = document.getElementById('importConfirmNote');
    var btn = this;
    btn.disabled = true;
    try{
      var res = await authFetch('/api/students/import/commit', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ import_token: pendingImportToken, duplicate_strategy: dupStrategy })
      });
      var result = await res.json();
      if (!result.ok){
        note.style.color = 'var(--absent)';
        note.textContent = result.error || 'Something went wrong.';
        return;
      }
      var parts = [result.added + ' added'];
      if (result.updated) parts.push(result.updated + ' replaced');
      if (result.skipped) parts.push(result.skipped + ' skipped');
      document.getElementById('importDoneMessage').textContent = parts.join(', ') + '.';
      showImportStep('importStepDone');
    } catch (err){
      note.style.color = 'var(--absent)';
      note.textContent = 'Could not reach the server.';
    } finally {
      btn.disabled = false;
    }
  });
})();
</script>
""",
    title="Dashboard — AAR",
)


@app.get("/", response_class=HTMLResponse)
async def landing() -> str:
    return LANDING_PAGE


@app.get("/signin", response_class=HTMLResponse)
async def signin_page() -> str:
    return SIGNIN_PAGE


@app.get("/signup", response_class=HTMLResponse)
async def signup_page() -> str:
    return SIGNUP_PAGE


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page() -> str:
    return DASHBOARD_PAGE


@app.get("/health")
async def health() -> dict:
    return {"ok": True, "time": datetime.now(timezone.utc).isoformat()}


# ---------- Auth ----------
def _do_signup(email: str, password: str, profile: dict) -> None:
    auth_result = supabase.auth.sign_up({"email": email, "password": password})
    if auth_result.user is None:
        raise ValueError("Sign up failed")
    # returning="minimal" skips the post-insert read-back, which would
    # otherwise need its own SELECT policy (and fail before the user has
    # a session, if email confirmation is still required).
    supabase.table(PROFILES_TABLE).insert(
        {"id": auth_result.user.id, **profile}, returning="minimal"
    ).execute()


@app.post("/api/signup")
async def api_signup(payload: dict):
    email = str(payload.get("email", "")).strip()
    password = str(payload.get("password", ""))
    first_name = str(payload.get("first_name", "")).strip()
    last_name = str(payload.get("last_name", "")).strip()
    institute_name = str(payload.get("institute_name", "")).strip()
    institute_branch_name = str(payload.get("institute_branch_name", "")).strip()
    mobile_number = str(payload.get("mobile_number", "")).strip()
    student_count_range = str(payload.get("student_count_range", "")).strip()

    if (
        not email or len(password) < 8 or not first_name or not last_name
        or not institute_name or not mobile_number
        or student_count_range not in STUDENT_COUNT_RANGES
    ):
        return JSONResponse({"ok": False, "error": "Please fill in all required fields."}, status_code=400)

    profile = {
        "first_name": first_name,
        "last_name": last_name,
        "institute_name": institute_name,
        "institute_branch_name": institute_branch_name or None,
        "email": email,
        "mobile_number": mobile_number,
        "student_count_range": student_count_range,
    }

    try:
        await asyncio.to_thread(_do_signup, email, password, profile)
    except AuthApiError as e:
        return JSONResponse({"ok": False, "error": e.message}, status_code=400)
    except Exception as e:
        print(f"[signup] unexpected error: {e!r}")
        return JSONResponse({"ok": False, "error": f"Something went wrong: {e}"}, status_code=400)

    return {"ok": True}


@app.post("/api/signin")
async def api_signin(payload: dict):
    email = str(payload.get("email", "")).strip()
    password = str(payload.get("password", ""))

    if not email or not password:
        return JSONResponse({"ok": False, "reason": "invalid_credentials"}, status_code=400)

    def _do_signin():
        return supabase.auth.sign_in_with_password({"email": email, "password": password})

    try:
        result = await asyncio.to_thread(_do_signin)
    except AuthApiError as e:
        reason = "not_verified" if e.code == "email_not_confirmed" else "invalid_credentials"
        return JSONResponse({"ok": False, "reason": reason}, status_code=401)
    except Exception as e:
        print(f"[signin] unexpected error: {e!r}")
        return JSONResponse({"ok": False, "reason": "invalid_credentials"}, status_code=401)

    if result.session is None:
        return JSONResponse({"ok": False, "reason": "not_verified"}, status_code=401)

    return {
        "ok": True,
        "access_token": result.session.access_token,
        "refresh_token": result.session.refresh_token,
    }


@app.post("/api/forgot-password")
async def api_forgot_password(payload: dict):
    email = str(payload.get("email", "")).strip()
    if email:
        try:
            await asyncio.to_thread(supabase.auth.reset_password_for_email, email)
        except Exception:
            pass  # never reveal whether an email is registered
    return {"ok": True}


@app.post("/api/refresh")
async def api_refresh(payload: dict):
    refresh_token = str(payload.get("refresh_token", "")).strip()
    if not refresh_token:
        return JSONResponse({"ok": False}, status_code=400)

    try:
        result = await asyncio.to_thread(supabase.auth.refresh_session, refresh_token)
    except Exception:
        return JSONResponse({"ok": False}, status_code=401)

    if result.session is None:
        return JSONResponse({"ok": False}, status_code=401)

    return {
        "ok": True,
        "access_token": result.session.access_token,
        "refresh_token": result.session.refresh_token,
    }


# ---------- Dashboard: stats ----------
@app.get("/api/dashboard/stats")
async def dashboard_stats(institute_id: str = Depends(get_institute_id)):
    def _query():
        count_result = (
            supabase.table(STUDENTS_TABLE)
            .select("aaruid", count="exact")
            .eq("institute_id", institute_id)
            .execute()
        )
        classes_result = (
            supabase.table(STUDENTS_TABLE).select("class").eq("institute_id", institute_id).execute()
        )
        distinct_classes = {r["class"] for r in classes_result.data if r.get("class")}
        return count_result.count or 0, len(distinct_classes)

    students_registered, classes_count = await asyncio.to_thread(_query)
    return {
        "ok": True,
        "devices_online": 0,  # no device registry/heartbeat exists yet
        "students_registered": students_registered,
        "classes_count": classes_count,
    }


# ---------- Dashboard: review gallery ----------
# Reference images stored in a public Supabase Storage bucket, one file per
# item (e.g. "ERASER.jpg"). Not institute-scoped - it's a shared catalog.
@app.get("/api/gallery")
async def list_gallery_images(institute_id: str = Depends(get_institute_id)):
    def _list():
        files = supabase.storage.from_(IMAGES_BUCKET).list()
        images = []
        for f in files:
            filename = f.get("name", "")
            if not filename or "." not in filename:
                continue
            display_name = filename.rsplit(".", 1)[0].upper()
            url = supabase.storage.from_(IMAGES_BUCKET).get_public_url(filename)
            images.append({"name": display_name, "url": url})
        images.sort(key=lambda item: item["name"])
        return images

    try:
        images = await asyncio.to_thread(_list)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True, "images": images}


# ---------- Dashboard: student management ----------
@app.get("/api/students")
async def list_students(institute_id: str = Depends(get_institute_id)):
    def _query():
        result = (
            supabase.table(STUDENTS_TABLE)
            .select("*")
            .eq("institute_id", institute_id)
            .order("name")
            .execute()
        )
        return result.data

    students = await asyncio.to_thread(_query)
    return {"ok": True, "students": students}


@app.post("/api/students")
async def create_student(payload: dict, institute_id: str = Depends(get_institute_id)):
    name = str(payload.get("name", "")).strip()
    rfid = str(payload.get("rfid", "")).strip()
    student_class = str(payload.get("class", "")).strip()
    section = str(payload.get("section", "")).strip()

    if not name or not rfid or not student_class or not section:
        return JSONResponse({"ok": False, "error": "All fields are required."}, status_code=400)

    aaruid = generate_aaruid()

    def _insert():
        supabase.table(STUDENTS_TABLE).insert(
            {
                "aaruid": aaruid,
                "institute_id": institute_id,
                "rfid": rfid,
                "name": name,
                "class": student_class,
                "section": section,
            },
            returning="minimal",
        ).execute()

    try:
        await asyncio.to_thread(_insert)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True, "aaruid": aaruid}


# ---------- Dashboard: device management ----------
@app.get("/api/devices")
async def list_devices(institute_id: str = Depends(get_institute_id)):
    def _query():
        result = (
            supabase.table(DEVICES_TABLE)
            .select("*")
            .eq("institute_id", institute_id)
            .order("name")
            .execute()
        )
        return result.data

    devices = await asyncio.to_thread(_query)
    return {"ok": True, "devices": devices}


@app.post("/api/devices")
async def create_device(payload: dict, institute_id: str = Depends(get_institute_id)):
    mac_address = str(payload.get("mac_address", "")).strip()
    name = str(payload.get("name", "")).strip()
    device_class = str(payload.get("class", "")).strip()
    section = str(payload.get("section", "")).strip()

    if not mac_address or not name or not device_class or not section:
        return JSONResponse({"ok": False, "error": "All fields are required."}, status_code=400)

    aaruid = generate_aaruid()

    def _insert():
        supabase.table(DEVICES_TABLE).insert(
            {
                "aaruid": aaruid,
                "institute_id": institute_id,
                "mac_address": mac_address,
                "name": name,
                "class": device_class,
                "section": section,
            },
            returning="minimal",
        ).execute()

    try:
        await asyncio.to_thread(_insert)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True, "aaruid": aaruid}


@app.patch("/api/devices/{aaruid}")
async def update_device(aaruid: str, payload: dict, institute_id: str = Depends(get_institute_id)):
    name = str(payload.get("name", "")).strip()
    device_class = str(payload.get("class", "")).strip()
    section = str(payload.get("section", "")).strip()

    if not name or not device_class or not section:
        return JSONResponse({"ok": False, "error": "All fields are required."}, status_code=400)

    def _update():
        supabase.table(DEVICES_TABLE).update(
            {"name": name, "class": device_class, "section": section},
            returning="minimal",
        ).eq("aaruid", aaruid).eq("institute_id", institute_id).execute()

    try:
        await asyncio.to_thread(_update)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True}


@app.delete("/api/devices/{aaruid}")
async def delete_device(aaruid: str, institute_id: str = Depends(get_institute_id)):
    def _delete():
        supabase.table(DEVICES_TABLE).delete(returning="minimal").eq("aaruid", aaruid).eq(
            "institute_id", institute_id
        ).execute()

    try:
        await asyncio.to_thread(_delete)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True}


@app.get("/api/students/template")
async def download_student_template():
    def _build() -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Name", "RFID", "Class", "Section"])
        ws.append(["John Doe", "AA:BB:CC:DD", "8", "A"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await asyncio.to_thread(_build)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=aar_student_template.xlsx"},
    )


REQUIRED_IMPORT_COLUMNS = ["name", "rfid", "class", "section"]


@app.post("/api/students/import/analyze")
async def analyze_student_import(
    file: UploadFile = File(...), institute_id: str = Depends(get_institute_id)
):
    content = await file.read()

    def _parse() -> list[dict]:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        header = [str(c.value).strip().lower() if c.value else "" for c in header_row]
        col_index = {h: i for i, h in enumerate(header)}
        missing = [c for c in REQUIRED_IMPORT_COLUMNS if c not in col_index]
        if missing:
            raise ValueError(f"Missing column(s): {', '.join(missing)}")

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(
                {
                    "name": str(row[col_index["name"]] or "").strip(),
                    "rfid": str(row[col_index["rfid"]] or "").strip(),
                    "class": str(row[col_index["class"]] or "").strip(),
                    "section": str(row[col_index["section"]] or "").strip(),
                }
            )
        return [r for r in rows if r["name"] and r["rfid"]]

    try:
        rows = await asyncio.to_thread(_parse)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Could not read file: {e}"}, status_code=400)

    if not rows:
        return JSONResponse(
            {"ok": False, "error": "No valid student rows found in the file."}, status_code=400
        )

    def _existing_rfids() -> set[str]:
        result = (
            supabase.table(STUDENTS_TABLE).select("rfid").eq("institute_id", institute_id).execute()
        )
        return {r["rfid"] for r in result.data}

    existing = await asyncio.to_thread(_existing_rfids)
    duplicates = [r for r in rows if r["rfid"] in existing]

    import_token = uuid.uuid4().hex
    PENDING_IMPORTS[import_token] = {"institute_id": institute_id, "rows": rows}

    return {
        "ok": True,
        "import_token": import_token,
        "total": len(rows),
        "new_count": len(rows) - len(duplicates),
        "duplicate_count": len(duplicates),
    }


@app.post("/api/students/import/commit")
async def commit_student_import(payload: dict, institute_id: str = Depends(get_institute_id)):
    import_token = str(payload.get("import_token", ""))
    strategy = str(payload.get("duplicate_strategy", "skip"))

    pending = PENDING_IMPORTS.pop(import_token, None)
    if not pending or pending["institute_id"] != institute_id:
        return JSONResponse(
            {"ok": False, "error": "Import session expired. Please re-upload the file."},
            status_code=400,
        )

    rows = pending["rows"]

    def _commit() -> dict:
        existing_result = (
            supabase.table(STUDENTS_TABLE)
            .select("aaruid,rfid")
            .eq("institute_id", institute_id)
            .execute()
        )
        existing_by_rfid = {r["rfid"]: r["aaruid"] for r in existing_result.data}

        to_insert = []
        to_update = []
        skipped = 0
        for row in rows:
            existing_aaruid = existing_by_rfid.get(row["rfid"])
            if existing_aaruid:
                if strategy == "replace":
                    to_update.append((existing_aaruid, row))
                else:
                    skipped += 1
                continue
            to_insert.append(
                {
                    "aaruid": generate_aaruid(),
                    "institute_id": institute_id,
                    "rfid": row["rfid"],
                    "name": row["name"],
                    "class": row["class"],
                    "section": row["section"],
                }
            )

        if to_insert:
            supabase.table(STUDENTS_TABLE).insert(to_insert, returning="minimal").execute()
        for aaruid, row in to_update:
            supabase.table(STUDENTS_TABLE).update(
                {"name": row["name"], "class": row["class"], "section": row["section"]},
                returning="minimal",
            ).eq("aaruid", aaruid).execute()

        return {"added": len(to_insert), "updated": len(to_update), "skipped": skipped}

    try:
        summary = await asyncio.to_thread(_commit)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

    return {"ok": True, **summary}


# ---------- Student sync helpers (ESP32 RFID scan protocol) ----------
# Plain synchronous Supabase client calls, run off the event loop thread via
# asyncio.to_thread so one slow request doesn't stall other connections.
def _get_all_students() -> list[dict]:
    result = supabase.table(STUDENTS_TABLE).select("*").execute()
    return result.data


def _upsert_student(uid: str, name: str, status: str) -> None:
    supabase.table(STUDENTS_TABLE).upsert({"uid": uid, "name": name, "status": status}).execute()


def _find_student(uid: str) -> dict | None:
    result = supabase.table(STUDENTS_TABLE).select("*").eq("uid", uid).execute()
    return result.data[0] if result.data else None


def _mark_present(uid: str) -> None:
    supabase.table(STUDENTS_TABLE).update({"status": "present"}).eq("uid", uid).execute()


# ---------- WebSocket endpoint ----------
# Devices (or a future dashboard) connect here. Message protocol is small
# JSON envelopes: {"type": "...", ...}.
#
#   Client -> server
#     {"type": "scan", "uid": "AA:BB:CC:DD"}
#       A card was tapped on a reader. Server looks the uid up, marks the
#       student present if found, and replies with a "scan_result".
#     {"type": "sync_students", "students": [{"uid","name","status"}, ...]}
#       Bulk push (e.g. a device catching up after being offline, or an
#       admin tool seeding data). Each entry is upserted.
#
#   Server -> client
#     {"type": "roster", "students": [...]}
#       Sent right after connecting, and again after anything changes.
#     {"type": "scan_result", "uid", "matched", "name"}
#       Reply to a "scan" message.
@app.websocket("/ws")
async def ws_endpoint(websocket: WebSocket) -> None:
    await registry.register(websocket)
    try:
        students = await asyncio.to_thread(_get_all_students)
        await websocket.send_text(json.dumps({"type": "roster", "students": students}))

        while True:
            raw = await websocket.receive_text()
            try:
                message = json.loads(raw)
            except ValueError:
                continue

            msg_type = message.get("type")

            if msg_type == "scan":
                uid = str(message.get("uid", "")).strip()
                if not uid:
                    continue
                student = await asyncio.to_thread(_find_student, uid)
                matched = student is not None
                name = student["name"] if student else ""
                if student:
                    await asyncio.to_thread(_mark_present, uid)
                    students = await asyncio.to_thread(_get_all_students)

                await websocket.send_text(json.dumps({
                    "type": "scan_result", "uid": uid, "matched": matched, "name": name,
                }))
                if matched:
                    await registry.broadcast({"type": "roster", "students": students}, exclude=websocket)

            elif msg_type == "sync_students":
                incoming = message.get("students", [])
                if not isinstance(incoming, list):
                    continue
                for item in incoming:
                    uid = str(item.get("uid", "")).strip()
                    name = str(item.get("name", "")).strip()
                    status = str(item.get("status", "absent")).strip()
                    if uid and name:
                        await asyncio.to_thread(_upsert_student, uid, name, status)
                students = await asyncio.to_thread(_get_all_students)
                await registry.broadcast({"type": "roster", "students": students})

    except WebSocketDisconnect:
        pass
    finally:
        registry.unregister(websocket)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
